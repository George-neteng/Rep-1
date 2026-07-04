"""report.py — выгрузка результата запроса в PDF (reportlab) или Excel (openpyxl)."""

import os
import json
from database import get_record

BASE = os.path.dirname(os.path.abspath(__file__))
REP_DIR = os.path.join(BASE, "static", "reports")
os.makedirs(REP_DIR, exist_ok=True)

# Кириллица в putText OpenCV не работает, а в reportlab нужен Unicode-шрифт.
_DEJAVU_PATHS = [
    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    "/usr/share/fonts/dejavu/DejaVuSans.ttf",
    "/usr/share/fonts/truetype/freefont/FreeSans.ttf",
]


def find_record(rec_id):
    return get_record(rec_id)


def _fmt(v):
    return json.dumps(v, ensure_ascii=False) if isinstance(v, (list, dict)) else v


def build_pdf(rec):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.pdfgen import canvas
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.utils import ImageReader

    font = "Helvetica"
    for p in _DEJAVU_PATHS:
        if os.path.exists(p):
            pdfmetrics.registerFont(TTFont("DejaVu", p))
            font = "DejaVu"
            break

    path = os.path.join(REP_DIR, f"report_{rec['id']}.pdf")
    c = canvas.Canvas(path, pagesize=A4)
    W, H = A4
    y = H - 2 * cm

    c.setFont(font, 16)
    c.drawString(2 * cm, y, "Отчёт: анализ положения мяча")
    y -= 1 * cm

    c.setFont(font, 10)
    for line in [
        f"ID запроса: {rec['id']}",
        f"Время: {rec['timestamp']}",
        f"Файл: {rec['filename']}  ({rec['media']})",
        f"Время обработки: {rec['elapsed']} с",
    ]:
        c.drawString(2 * cm, y, line)
        y -= 0.6 * cm

    y -= 0.3 * cm
    c.setFont(font, 12)
    c.drawString(2 * cm, y, "Статистика:")
    y -= 0.7 * cm
    c.setFont(font, 10)
    for k, v in rec["stats"].items():
        c.drawString(2.3 * cm, y, f"{k}: {_fmt(v)}")
        y -= 0.5 * cm
        if y < 7 * cm:
            c.showPage()
            c.setFont(font, 10)
            y = H - 2 * cm

    img_path = os.path.join(BASE, rec["result_url"].lstrip("/"))
    if rec["media"] == "image" and os.path.exists(img_path):
        try:
            img = ImageReader(img_path)
            iw, ih = img.getSize()
            maxw = W - 4 * cm
            scale = min(maxw / iw, (y - 2 * cm) / ih)
            if scale > 0:
                c.drawImage(img, 2 * cm, y - ih * scale,
                            iw * scale, ih * scale, preserveAspectRatio=True)
        except Exception:
            pass

    c.showPage()
    c.save()
    return path


def build_xlsx(rec):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт"

    ws.append(["Параметр", "Значение"])
    ws["A1"].font = ws["B1"].font = Font(bold=True)

    for label, key in [
        ("ID запроса", "id"), ("Время", "timestamp"),
        ("Файл", "filename"), ("Тип", "media"),
        ("Время обработки, с", "elapsed"),
    ]:
        ws.append([label, rec[key]])

    ws.append([])
    ws.append(["Статистика", ""])
    ws[f"A{ws.max_row}"].font = Font(bold=True)
    for k, v in rec["stats"].items():
        ws.append([k, _fmt(v)])

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 42

    path = os.path.join(REP_DIR, f"report_{rec['id']}.xlsx")
    wb.save(path)
    return path
